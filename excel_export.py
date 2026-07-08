from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ExcelExport:

    @staticmethod
    def export(df, output_file):

        wb = Workbook()
        ws = wb.active
        ws.title = "Vouchere"

        # antet
        headers = list(df.columns)

        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col)
            c.value = h
            c.font = Font(bold=True)

        # date
        for row in df.itertuples(index=False):

            ws.append(list(row))

        # autofilter
        ws.auto_filter.ref = ws.dimensions

        # freeze first row
        ws.freeze_panes = "A2"

        # autosize coloane
        for column in ws.columns:

            length = 0

            letter = get_column_letter(column[0].column)

            for cell in column:

                try:
                    length = max(length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[letter].width = min(length + 2, 60)

        wb.save(output_file)
